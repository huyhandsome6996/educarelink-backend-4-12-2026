"""Admin statistics dashboard — Thống kê số liệu landing page.

Accessible only to staff users at /admin/thong-ke/.
Features:
  - Real-time overview cards (visits, surveys, trial signups, consult signups)
  - Detailed tables per data type with filtering
  - AI-powered data analysis (Gemini)
  - Excel export (openpyxl)
"""

import json
from datetime import timedelta

from django.conf import settings
from django.contrib.admin.views.decorators import staff_member_required
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.utils import timezone
from django.db.models import Count, Q

from .models import LandingPageVisit, LandingSurvey, LandingSignup


@staff_member_required
def admin_statistics_view(request):
    """Trang thống kê tổng hợp landing page."""
    now = timezone.now()
    today = now.replace(hour=0, minute=0, second=0, microsecond=0)
    week_ago = today - timedelta(days=7)
    month_ago = today - timedelta(days=30)

    # --- Tổng quan ---
    total_visits = LandingPageVisit.objects.count()
    visits_today = LandingPageVisit.objects.filter(visited_at__gte=today).count()
    visits_week = LandingPageVisit.objects.filter(visited_at__gte=week_ago).count()
    visits_month = LandingPageVisit.objects.filter(visited_at__gte=month_ago).count()

    total_surveys = LandingSurvey.objects.count()
    surveys_today = LandingSurvey.objects.filter(created_at__gte=today).count()
    surveys_cp = LandingSurvey.objects.filter(role='carepartner').count()
    surveys_ph = LandingSurvey.objects.filter(role='phu-huynh').count()

    total_signups = LandingSignup.objects.count()
    signups_today = LandingSignup.objects.filter(created_at__gte=today).count()
    trial_signups = LandingSignup.objects.filter(signup_type='dung-thu').count()
    consult_signups = LandingSignup.objects.filter(signup_type='tu-van').count()

    # --- Recent items for tables ---
    recent_visits = LandingPageVisit.objects.all()[:50]
    recent_surveys = LandingSurvey.objects.select_related().all()[:50]
    recent_signups = LandingSignup.objects.all()[:50]

    # --- Phân tích survey theo role ---
    survey_role_breakdown = list(
        LandingSurvey.objects.values('role').annotate(count=Count('id')).order_by('-count')
    )

    # --- Signup type breakdown ---
    signup_type_breakdown = list(
        LandingSignup.objects.values('signup_type').annotate(count=Count('id')).order_by('-count')
    )

    # --- Daily visits (last 14 days) for chart ---
    daily_visits = []
    for i in range(13, -1, -1):
        day = today - timedelta(days=i)
        day_end = day + timedelta(days=1)
        count = LandingPageVisit.objects.filter(
            visited_at__gte=day, visited_at__lt=day_end
        ).count()
        daily_visits.append({'date': day.strftime('%d/%m'), 'count': count})

    # --- Daily surveys (last 14 days) ---
    daily_surveys = []
    for i in range(13, -1, -1):
        day = today - timedelta(days=i)
        day_end = day + timedelta(days=1)
        count = LandingSurvey.objects.filter(
            created_at__gte=day, created_at__lt=day_end
        ).count()
        daily_surveys.append({'date': day.strftime('%d/%m'), 'count': count})

    # --- Daily signups (last 14 days) ---
    daily_signups = []
    for i in range(13, -1, -1):
        day = today - timedelta(days=i)
        day_end = day + timedelta(days=1)
        count = LandingSignup.objects.filter(
            created_at__gte=day, created_at__lt=day_end
        ).count()
        daily_signups.append({'date': day.strftime('%d/%m'), 'count': count})

    context = {
        'title': 'Thống kê số liệu — EduCareLink Landing Page',
        # Overview
        'total_visits': total_visits,
        'visits_today': visits_today,
        'visits_week': visits_week,
        'visits_month': visits_month,
        'total_surveys': total_surveys,
        'surveys_today': surveys_today,
        'surveys_cp': surveys_cp,
        'surveys_ph': surveys_ph,
        'total_signups': total_signups,
        'signups_today': signups_today,
        'trial_signups': trial_signups,
        'consult_signups': consult_signups,
        # Tables
        'recent_visits': recent_visits,
        'recent_surveys': recent_surveys,
        'recent_signups': recent_signups,
        # Breakdowns
        'survey_role_breakdown': survey_role_breakdown,
        'signup_type_breakdown': signup_type_breakdown,
        # Chart data
        'daily_visits': json.dumps(daily_visits),
        'daily_surveys': json.dumps(daily_surveys),
        'daily_signups': json.dumps(daily_signups),
        # Has AI key?
        'has_gemini': bool(settings.GEMINI_API_KEY),
    }
    return render(request, 'admin/thong_ke.html', context)


@staff_member_required
def admin_export_excel(request):
    """Xuất toàn bộ dữ liệu landing page ra file Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # --- Style ---
    header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='F26522', end_color='F26522', fill_type='solid')
    cell_font = Font(name='Arial', size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    wrap_align = Alignment(wrap_text=True, vertical='top')

    def style_header(ws, num_cols):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

    def style_cells(ws, num_cols, num_rows):
        for row in range(2, num_rows + 1):
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = cell_font
                cell.border = thin_border
                cell.alignment = wrap_align

    # ===== Sheet 1: Lượt truy cập =====
    ws1 = wb.active
    ws1.title = 'Lượt truy cập'
    visits = LandingPageVisit.objects.all().order_by('-visited_at')
    headers1 = ['ID', 'Session ID', 'IP', 'User Agent', 'Referrer', 'Thời gian']
    for i, h in enumerate(headers1, 1):
        ws1.cell(row=1, column=i, value=h)
    for idx, v in enumerate(visits, 2):
        ws1.cell(row=idx, column=1, value=v.id)
        ws1.cell(row=idx, column=2, value=v.session_id)
        ws1.cell(row=idx, column=3, value=v.ip_address or '')
        ws1.cell(row=idx, column=4, value=v.user_agent[:200])
        ws1.cell(row=idx, column=5, value=v.referrer or '')
        tz = timezone.get_current_timezone()
        ws1.cell(row=idx, column=6, value=v.visited_at.astimezone(tz).strftime('%d/%m/%Y %H:%M:%S'))
    style_header(ws1, len(headers1))
    style_cells(ws1, len(headers1), len(visits) + 1)
    for col, w in enumerate([6, 40, 18, 50, 40, 22], 1):
        ws1.column_dimensions[get_column_letter(col)].width = w

    # ===== Sheet 2: Khảo sát góp ý =====
    ws2 = wb.create_sheet('Khảo sát góp ý')
    surveys = LandingSurvey.objects.all().order_by('-created_at')
    headers2 = ['ID', 'Vai trò', 'Câu trả lời (JSON)', 'Góp ý', 'Email', 'IP', 'Thời gian']
    for i, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=i, value=h)
    for idx, s in enumerate(surveys, 2):
        ws2.cell(row=idx, column=1, value=s.id)
        ws2.cell(row=idx, column=2, value=s.get_role_display())
        ws2.cell(row=idx, column=3, value=json.dumps(s.role_answers, ensure_ascii=False) if s.role_answers else '')
        ws2.cell(row=idx, column=4, value=s.feedback)
        ws2.cell(row=idx, column=5, value=s.email or '')
        ws2.cell(row=idx, column=6, value=s.ip_address or '')
        tz = timezone.get_current_timezone()
        ws2.cell(row=idx, column=7, value=s.created_at.astimezone(tz).strftime('%d/%m/%Y %H:%M:%S'))
    style_header(ws2, len(headers2))
    style_cells(ws2, len(headers2), len(surveys) + 1)
    for col, w in enumerate([6, 18, 60, 50, 30, 18, 22], 1):
        ws2.column_dimensions[get_column_letter(col)].width = w

    # ===== Sheet 3: Đăng ký tư vấn/dùng thử =====
    ws3 = wb.create_sheet('Đăng ký')
    signups = LandingSignup.objects.all().order_by('-created_at')
    headers3 = ['ID', 'Họ tên', 'SĐT', 'Email', 'Vai trò', 'Loại đăng ký',
                'Khung giờ', 'Đồng ý dùng thử', 'Ghi chú', 'IP', 'Thời gian']
    for i, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=i, value=h)
    for idx, s in enumerate(signups, 2):
        ws3.cell(row=idx, column=1, value=s.id)
        ws3.cell(row=idx, column=2, value=s.full_name)
        ws3.cell(row=idx, column=3, value=s.phone)
        ws3.cell(row=idx, column=4, value=s.email)
        ws3.cell(row=idx, column=5, value=s.get_role_display())
        ws3.cell(row=idx, column=6, value=s.get_signup_type_display())
        ws3.cell(row=idx, column=7, value=s.get_preferred_time_slot_display() if s.preferred_time_slot else '')
        ws3.cell(row=idx, column=8, value='Có' if s.trial_consent else '')
        ws3.cell(row=idx, column=9, value=s.note)
        ws3.cell(row=idx, column=10, value=s.ip_address or '')
        tz = timezone.get_current_timezone()
        ws3.cell(row=idx, column=11, value=s.created_at.astimezone(tz).strftime('%d/%m/%Y %H:%M:%S'))
    style_header(ws3, len(headers3))
    style_cells(ws3, len(headers3), len(signups) + 1)
    for col, w in enumerate([6, 25, 16, 30, 18, 20, 24, 16, 40, 18, 22], 1):
        ws3.column_dimensions[get_column_letter(col)].width = w

    # ===== Response =====
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        'attachment; filename="educarelink-thong-ke-landing-page.xlsx"'
    )
    wb.save(response)
    return response


@staff_member_required
def admin_ai_analysis(request):
    """AI phân tích dữ liệu landing page (Gemini)."""
    if not settings.GEMINI_API_KEY:
        return JsonResponse({'error': 'Chưa cấu hình GEMINI_API_KEY.'}, status=500)

    import google.generativeai as genai
    genai.configure(api_key=settings.GEMINI_API_KEY)

    # Gather stats
    now = timezone.now()
    today = now.replace(hour=0, minute=0, second=0, microsecond=0)
    month_ago = today - timedelta(days=30)

    total_visits = LandingPageVisit.objects.count()
    visits_month = LandingPageVisit.objects.filter(visited_at__gte=month_ago).count()
    total_surveys = LandingSurvey.objects.count()
    surveys_cp = LandingSurvey.objects.filter(role='carepartner').count()
    surveys_ph = LandingSurvey.objects.filter(role='phu-huynh').count()
    total_signups = LandingSignup.objects.count()
    trial_count = LandingSignup.objects.filter(signup_type='dung-thu').count()
    consult_count = LandingSignup.objects.filter(signup_type='tu-van').count()

    # Get sample feedbacks
    recent_feedbacks = list(
        LandingSurvey.objects.exclude(feedback='').order_by('-created_at')[:10]
        .values_list('role', 'feedback', 'created_at')
    )
    feedback_text = '\n'.join(
        f'- [{r[0]}] {r[1][:200]} ({r[2].strftime("%d/%m/%Y")})'
        for r in recent_feedbacks
    ) if recent_feedbacks else 'Chưa có góp ý.'

    # Survey role_answers analysis
    cp_surveys = LandingSurvey.objects.filter(role='carepartner')[:20]
    ph_surveys = LandingSurvey.objects.filter(role='phu-huynh')[:20]
    cp_answers = [s.role_answers for s in cp_surveys if s.role_answers]
    ph_answers = [s.role_answers for s in ph_surveys if s.role_answers]

    prompt = f"""Bạn là chuyên gia phân tích dữ liệu cho cuộc thi khởi nghiệp.

Phân tích các số liệu sau từ landing page EduCareLink và đưa ra insight có giá trị
để làm báo cáo cho cuộc thi:

**TỔNG QUAN:**
- Tổng lượt truy cập (real, không bot): {total_visits}
- Lượt truy cập 30 ngày gần nhất: {visits_month}
- Tổng khảo sát góp ý: {total_surveys} (CarePartner: {surveys_cp}, Phụ huynh: {surveys_ph})
- Tổng đăng ký: {total_signups} (Dùng thử: {trial_count}, Tư vấn: {consult_count})

**MẪU GÓP Ý GẦN ĐÂY:**
{feedback_text}

**CÂU TRẢ LỜI CHI TIẾT (CarePartner - 20 mẫu mới nhất):**
{json.dumps(cp_answers, ensure_ascii=False, indent=2) if cp_answers else 'Chưa có'}

**CÂU TRẢ LỜI CHI TIẾT (Phụ huynh - 20 mẫu mới nhất):**
{json.dumps(ph_answers, ensure_ascii=False, indent=2) if ph_answers else 'Chưa có'}

Hãy phân tích:
1. Đánh giá tổng thể — số liệu này có tốt cho cuộc thi không?
2. Insight từ câu trả lời — nhu cầu chính của từng nhóm là gì?
3. Điểm mạnh/điểm yếu từ phản hồi
4. Gợi ý hành động cụ thể để tăng conversion rate
5. 3 số liệu highlight quan trọng nhất nên đưa vào báo cáo

Trả lời bằng tiếng Việt, ngắn gọn, súc tích, định dạng Markdown.
"""

    try:
        model = genai.GenerativeModel('gemini-2.0-flash')
        response = model.generate_content(prompt)
        return JsonResponse({'analysis': response.text})
    except Exception as e:
        return JsonResponse({'error': f'Lỗi AI: {str(e)}'}, status=500)
